import configparser
import logging

import openpyxl
from PIL import ImageGrab
from datetime import datetime
import time
import os
from selenium import webdriver
from selenium.common import TimeoutException
from selenium.webdriver.edge.service import Service as EdgeService, Service
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager


class GeneralUtils:
    def __init__(self, driver=None):
        self.driver = driver # driver placeholder

    def open_in_browser(self, url):
        from selenium import webdriver
        from selenium.webdriver.edge.options import Options

        edge_options = Options()
        edge_options.add_argument("--start-maximized")
        edge_options.add_argument("--disable-web-security")
        edge_options.add_argument("--no-first-run")
        edge_options.add_argument("--disable-features=DefaultBrowserSettingEnabled")
        self.driver = webdriver.Edge(options=edge_options)
        self.driver.get(url)

    def get_driver(self):
        return self.driver

    def close_browser(self):
        if self.driver:
            self.driver.quit()
        else:
            logging.warning("Driver is not initialized.")


def read_ini_config():
    config = configparser.ConfigParser()
    # file_path = os.path.abspath(os.path.join(os.curdir,'..', '..',  'Configurations', 'config.ini'))
    # Get the directory of the current script
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Construct the full path to the config.ini file
    relative_path  = os.path.join(script_dir,'..', 'Configurations', 'config.ini')
    file_path = os.path.abspath(relative_path)
    # logger.info(f"Absolute path to config.ini file: {file_path}")
    config_dict = {}
    try:
        # Read the INI file
        config.read(file_path)
        for section in config.sections():
            config_dict[section] = {}
            for key, value in config.items(section):
                config_dict[section][key] = value
    except configparser.Error as e:
        print(f"Error reading INI file: {e}")
    return config_dict.get('commonInfo', {})

def delete_old_screenshots():
    """Delete all screenshots in the specified folder."""
    folder_path = os.path.abspath(os.path.join(os.curdir, 'ScreenShots'))
    print(f"Directory path: {folder_path}")

    if os.path.exists(folder_path):
        print(f"Directory exists: {folder_path}")
        for file_name in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file_name)
            print(f"Attempting to delete: {file_path}")

            if os.path.isfile(file_path):
                try:
                    os.remove(file_path)
                    print(f"Deleted old screenshot: {file_path}")
                except Exception as e:
                    print(f"Error deleting file {file_path}: {e}")
    else:
        print(f"Directory not found: {folder_path}")


def capture_and_save_screenshot(file_name="ScreenShot"):
    folder_path = os.path.abspath(os.path.join(os.curdir, 'ScreenShots'))
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    screenshot = ImageGrab.grab()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(folder_path, f"{file_name}_{timestamp}.png")
    screenshot.save(file_path)
    print(f"Screenshot saved at: {file_path}")
    return file_path

def get_excel_test_data(file_name, sheet_name, test_id):
        xl_file = openpyxl.load_workbook(file_name)
        sheet = xl_file[sheet_name]
        headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == test_id:
                data = {header.value: value for header, value in zip(headers, row)}
                return data
        print(f"Error: Test ID '{test_id}' not found in sheet '{sheet_name}'.")

logging.basicConfig(level=logging.INFO)

def default_wait(driver, locator_type=None, locator_value=None, condition=None, timeout=40, multiple=False, custom_lambda=None):
    """
    Waits for a single element, multiple elements, or a custom condition using WebDriverWait.
    """
    wait = WebDriverWait(driver, timeout)

    if custom_lambda:
        try:
            logging.info("Waiting using custom lambda condition")
            return wait.until(custom_lambda)
        except TimeoutException as e:
            logging.error(f"Timeout: Custom condition not met in {timeout} seconds: {e}")
            raise
        except Exception as e:
            logging.error(f"Error during custom lambda wait: {e}")
            raise

    condition = condition or (EC.presence_of_all_elements_located if multiple else EC.presence_of_element_located)

    try:
        if multiple:
            elements = wait.until(condition((locator_type, locator_value)))
            logging.info(f"Found {len(elements)} elements: {[e.text for e in elements]}")
            return elements
        else:
            element = wait.until(condition((locator_type, locator_value)))
            logging.info("Element found and ready.")
            return element
    except TimeoutException as e:
        logging.error(f"Timeout: Element not found within {timeout} seconds: {e}")
        raise
    except Exception as e:
        logging.error(f"Error while waiting for element: {e}")
        raise

def default_wait_for_both(driver, by, value, timeout=30):
    try:
        # Check if the value corresponds to multiple elements
        elements = driver.find_elements(by, value)
        if len(elements) > 1:
            return elements
        else:
            single_element = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))
            print(single_element)
            return single_element
    except Exception as e:
        print(f"An error occurred while waiting for the element: {e}")

def _convert_date(date_obj):
    date_obj = datetime.strptime(date_obj, '%Y-%m-%d')
    return date_obj.strftime('%m/%d/%Y') # Convert datetime object to string

def sleep(seconds):
    time.sleep(seconds)
